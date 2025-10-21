import io
from pathlib import Path
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="Select Sheets", page_icon="📄", layout="centered")
st.title("📄 Extract Selected Sheets from Excel")

# ---------------------------
# Helpers for stable uploads
# ---------------------------
def _persist_upload_to_state(file, slot_key_bytes: str, slot_key_name: str):
    """
    Read uploaded file into bytes and store in session_state so the user
    can upload files one-by-one without losing the first upload.
    """
    if file is None:
        return
    st.session_state[slot_key_bytes] = file.getvalue()
    st.session_state[slot_key_name]  = file.name

def _get_persisted_file(slot_key_bytes: str, slot_key_name: str):
    """
    Return (name, bytes) from session_state if present, else (None, None).
    """
    name = st.session_state.get(slot_key_name)
    b    = st.session_state.get(slot_key_bytes)
    return name, b

# Init state keys
for k in ["excel_bytes", "excel_name", "csv_bytes", "csv_name"]:
    st.session_state.setdefault(k, None)

col1, col2 = st.columns(2)
with col1:
    xlsx_file = st.file_uploader(
        "Upload Excel (.xlsx/.xlsm)",
        type=["xlsx", "xlsm"],
        key="uploader_excel",
        help="You can upload this first, then upload the CSV next.",
    )
    _persist_upload_to_state(xlsx_file, "excel_bytes", "excel_name")
    if st.session_state["excel_name"]:
        st.caption(f"Excel: {st.session_state['excel_name']} ✅")

with col2:
    csv_file = st.file_uploader(
        "Upload CSV with 'SheetName' column",
        type=["csv"],
        key="uploader_csv",
        help="You can upload this after the Excel; both will be kept in memory.",
    )
    _persist_upload_to_state(csv_file, "csv_bytes", "csv_name")
    if st.session_state["csv_name"]:
        st.caption(f"CSV: {st.session_state['csv_name']} ✅")

# Clear buttons (optional, handy while testing)
cc1, cc2 = st.columns(2)
with cc1:
    if st.button("Clear Excel"):
        st.session_state["excel_bytes"] = None
        st.session_state["excel_name"] = None
with cc2:
    if st.button("Clear CSV"):
        st.session_state["csv_bytes"] = None
        st.session_state["csv_name"] = None

# Resolve current effective inputs
excel_name, excel_bytes = _get_persisted_file("excel_bytes", "excel_name")
csv_name,   csv_bytes   = _get_persisted_file("csv_bytes", "csv_name")

# Show current detection
st.markdown("**Detected**")
st.write("• Excel:", excel_name or "—")
st.write("• CSV:", csv_name or "—")

# Output mode
mode = st.radio(
    "Choose output",
    ["Filtered workbook", "Summary sheet", "Both"],
    index=0,
    key="mode_radio",
)

# Ready when both are present
ready = (excel_bytes is not None) and (csv_bytes is not None)

# ---------------------------
# Build logic
# ---------------------------
def build_outputs(excel_name: str, excel_bytes: bytes, csv_bytes: bytes, mode: str):
    # --- Read CSV for desired tabs ---
    wanted_df = pd.read_csv(io.BytesIO(csv_bytes), dtype=str)
    if "SheetName" not in wanted_df.columns:
        st.error("CSV must contain a column named 'SheetName'.")
        st.stop()
    wanted = wanted_df["SheetName"].dropna().map(str.strip).tolist()
    if not wanted:
        st.error("No sheet names found in CSV.")
        st.stop()

    # Macro detection & output ext
    keep_vba = Path(excel_name).suffix.lower() == ".xlsm"
    filtered_ext = ".xlsm" if keep_vba else ".xlsx"

    # Load twice: preserve + values
    wb = load_workbook(io.BytesIO(excel_bytes), data_only=False, keep_vba=keep_vba)
    wb_vals = load_workbook(io.BytesIO(excel_bytes), data_only=True)

    source_sheetnames = wb.sheetnames
    wanted_set = set(wanted)
    matched = [s for s in wanted if s in source_sheetnames]
    missing = [s for s in wanted if s not in source_sheetnames]
    if not matched:
        st.error("No matching sheets found in the workbook.")
        st.stop()

    # A) Filtered workbook
    def build_filtered_workbook_bytes():
        for name in list(wb.sheetnames):
            if name not in wanted_set:
                wb.remove(wb[name])
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return buf, f"SelectedSheets{filtered_ext}"

    # B) Summary sheet
    def build_summary_bytes():
        TARGET = "total net asset value after ic fall"

        def norm(s):
            if s is None:
                return ""
            return " ".join(str(s).strip().lower().split())

        rows = []
        for tab in matched:
            ws = wb_vals[tab]

            # Name B4
            try:
                name_val = ws["B4"].value
            except Exception:
                name_val = None

            # Cash C27
            try:
                cash_val = ws["C27"].value
            except Exception:
                cash_val = None

            # NAV: find label in col A, get value from col B same row
            nav_val = None
            try:
                for a_cell in ws["A"]:
                    label = norm(a_cell.value)
                    if label and (label == norm(TARGET) or norm(TARGET) in label):
                        nav_val = ws.cell(row=a_cell.row, column=2).value  # Column B
                        break
            except Exception:
                nav_val = None

            rows.append({"Name": name_val, "NAV": nav_val, "Cash": cash_val})

        df = pd.DataFrame(rows)
        for col in ["NAV", "Cash"]:
            df[col] = pd.to_numeric(df[col], errors="coerce")

        out = io.BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as xw:
            df.to_excel(xw, index=False, sheet_name="Summary")
        out.seek(0)
        return out, "Summary.xlsx", missing

    # Build per mode
    if mode in ("Filtered workbook", "Both"):
        fbuf, fname = build_filtered_workbook_bytes()
        st.success(f"✅ Built {fname}")
        st.download_button("Download filtered workbook", data=fbuf, file_name=fname, key="dl_filtered")

    if mode in ("Summary sheet", "Both"):
        sbuf, sname, missing = build_summary_bytes()
        st.success(f"✅ Built {sname}")
        st.download_button("Download summary sheet", data=sbuf, file_name=sname, key="dl_summary")
        if missing:
            st.warning("Missing sheets (not found in workbook): " + ", ".join(missing))

# Button
btn = st.button("Build file(s)", disabled=not ready, key="build_btn")

if btn:
    try:
        build_outputs(excel_name, excel_bytes, csv_bytes, mode)
    except Exception as e:
        st.error(f"Error: {e}")

# Debug footer (optional, comment out if not needed)
with st.expander("Runtime info (debug)"):
    import sys
    st.write("Python:", sys.version)
    try:
        import openpyxl
        st.write("openpyxl:", openpyxl.__version__)
    except Exception as e:
        st.write("openpyxl import error:", e)
    try:
        import pandas as pd as _pd_tmp  # noqa
        st.write("pandas:", pd.__version__)
    except Exception as e:
        st.write("pandas import error:", e)

